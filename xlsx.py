import re
from constants import *
from sqlalchemy import Sequence
from table_definitions import *
from openpyxl import load_workbook, Workbook



'''
=====================================================================================
| Custom Exception that will be raised if there's a problem with the XLSX file.     |
=====================================================================================
'''
class XLSXException(Exception):
    pass

'''
=====================================================================================
| Helper class for managing reading and writing to XLSX files.                      |
=====================================================================================
'''
class XLSXHandler():

    def __init__(self, file_path: str, sheet_name: str, db: Session):
        self.file_path: str = file_path
        self.doc_values: Workbook = load_workbook(file_path, data_only=True) # One workbook for the values...
        self.doc_formulas: Workbook = load_workbook(file_path, data_only=False) # ... and one workbook for the formulas.
        self.sheet_name: str = sheet_name
        self.rows: list = self.read_rows()
        self.db: Session = db



    '''
    =====================================================================================
    | Reads all rows from the worksheet and returns it as a list of lists.              |
    =====================================================================================
    '''
    def read_rows(self) -> list:
        rows = [] # TODO: CONVERT TO HASHTABLE

        # Get the worksheet with values (not the formulas) for reading.
        try:
            sheet = self.doc_values[self.sheet_name]
        except:
            raise XLSXException(f"Worksheet \"{self.sheet_name}\" not found.")

        for row in sheet.iter_rows(min_row=2):
            row_data = []
            for cell in row:
                row_data.append(cell.value)
            
            rows.append(row_data)
        
        return rows
    


    '''
    =====================================================================================
    | Match all techniques in the change database against the XLSX file. If a technique |
    | from the database is found in the XLSX file, import it's values. If a technique   |
    | is not found, all values are set to default.                                      |
    =====================================================================================
    '''
    def import_xlsx(self, changes: Sequence[MITREChange]) -> None:
        mitre_ids = [row[COL_MITREID] for row in self.rows]

        for c in changes:
            # If a technique is not in the xlsx file, set everything to their default value.
            if c.mitre_id not in mitre_ids:
                # Client scores.
                c.client_criticality, c.client_criticality_sum = 0, 0
                c.client_evaluation_status = "not evaluated"
                c.client_reasoning, c.client_measures = "", ""

                # Infrastructure scores.
                c.infra_criticality, c.infra_criticality_sum = 0, 0
                c.infra_evaluation_status = "not evaluated"
                c.infra_reasoning, c.infra_measures = "", ""

                # Service scores.
                c.service_criticality, c.service_criticality_sum = 0, 0
                c.service_evaluation_status = "not evaluated"
                c.infra_reasoning, c.infra_measures = "", ""

                c.confidentiality, c.integrity, c.availability = False, False, False
                continue

            # Else, search for the row and import the values into the DB.
            for row in self.rows:
                if row[COL_MITREID] == c.mitre_id:
                    # Client scores.
                    c.client_criticality = 0 if row[COL_CLIENT_CRITICALITY] == "n.a." else row[COL_CLIENT_CRITICALITY]
                    c.client_criticality_sum = 0 if row[COL_CLIENT_CRITICALITY_SUM] == "n.a." else row[COL_CLIENT_CRITICALITY_SUM]
                    c.client_evaluation_status = row[COL_CLIENT_EVALUATION_STATUS]
                    c.client_reasoning = row[COL_CLIENT_REASONING]
                    c.client_measures = row[COL_CLIENT_MEASURES]

                    # Infrastructure scores.
                    c.infra_criticality = 0 if row[COL_INFRASTRUCTURE_CRITICALITY] == "n.a." else row[COL_INFRASTRUCTURE_CRITICALITY]
                    c.infra_criticality_sum = 0 if row[COL_INFRASTRUCTURE_CRITICALITY_SUM] == "n.a." else row[COL_INFRASTRUCTURE_CRITICALITY_SUM]
                    c.infra_evaluation_status = row[COL_INFRASTRUCTURE_EVALUATION_STATUS]
                    c.infra_reasoning = row[COL_INFRASTRUCTURE_REASONING]
                    c.infra_measures = row[COL_INFRASTRUCTURE_MEASURES]

                    # Service scores.
                    c.service_criticality = 0 if row[COL_SERVICE_CRITICALITY] == "n.a." else row[COL_SERVICE_CRITICALITY]
                    c.service_criticality_sum = 0 if row[COL_SERVICE_CRITICALITY_SUM] == "n.a." else row[COL_SERVICE_CRITICALITY_SUM]
                    c.service_evaluation_status = row[COL_SERVICE_EVALUATION_STATUS]
                    c.service_reasoning = row[COL_SERVICE_REASONING]
                    c.service_measures = row[COL_SERVICE_MEASURES]

                    # If "x" then True, else False.
                    c.confidentiality = row[COL_CONFIDENTIALITY] == "x"
                    c.integrity = row[COL_INTEGRITY] == "x"
                    c.availability = row[COL_AVAILABILITY] == "x"
            
        self.db.commit()


    '''
    =====================================================================================
    | Small helper function that aligns a change with the headers of the XLSX file.     |
    =====================================================================================
    '''
    @staticmethod
    def create_row(change: MITREChange) -> list:
        return [None, change.mitre_id, change.tactics, change.technique, change.sub_technique,
            change.client_criticality, change.infra_criticality, change.service_criticality,
            "x" if change.confidentiality else None, "x" if change.integrity else None,
            "x" if change.availability else None, change.client_criticality_sum,
            None, None, None, None, change.infra_criticality_sum, None, None, None, None,
            change.service_criticality_sum, None, None, None, None
        ]
    

    
    '''
    =====================================================================================
    | This function detects if a row in the worksheet is completely empty.              |
    =====================================================================================
    '''
    def is_empty_row(self, row) -> bool:
        for cell in row:
            if cell.value not in (None, ""):
                return False
            
        return True


    '''
    =====================================================================================
    | This function exports the SQLite database and all changes made during the upgrade |
    | back into the XLSX file and saves it on the local disk.                           |
    | Additions are added in the bottom of the Worksheet. Sorting is currently not      |
    | supported.                                                                        |
    =====================================================================================
    '''
    def export_xlsx(self, file_path: str, changes: Sequence[MITREChange]):
        sheet = self.doc_formulas[self.sheet_name]

        for c in changes:
            # Append to the sheet if it's a new addition.
            if c.change_category == "additions":
                sheet.append(self.create_row(c))
                continue
            
            # If not a new addition, change the row.
            # Iterate over all rows to find the correct one.
            for row in sheet.iter_rows(min_row=2):
                # Get the MITREID. If it's empty use an empty string by default (or else it will be None and a TypeError happens with regex).
                mitre_id = row[COL_MITREID].value or ""

                # Extract with regex pattern.
                pattern = r'HYPERLINK\([^,;]+[,;]\s*"([^"]+)"\)'
                match = re.search(pattern, mitre_id)
                mitre_id = match.group(1) if match else None

                # If MITRE ID matches, change the row.
                if mitre_id == c.mitre_id:
                    # Client Scores.
                    # Criticality Sums don't need to be exported since they are calculated by a formula in the file.
                    row[COL_CLIENT_CRITICALITY].value = c.client_criticality
                    row[COL_CLIENT_EVALUATION_STATUS].value = c.client_evaluation_status
                    row[COL_CLIENT_REASONING].value = c.client_reasoning
                    row[COL_CLIENT_MEASURES].value = c.client_measures

                    # Infrastructure Scores.
                    row[COL_INFRASTRUCTURE_CRITICALITY].value = c.infra_criticality
                    row[COL_INFRASTRUCTURE_EVALUATION_STATUS].value = c.infra_evaluation_status
                    row[COL_INFRASTRUCTURE_REASONING].value = c.infra_reasoning
                    row[COL_INFRASTRUCTURE_MEASURES].value = c.infra_measures

                    # Service Scores.
                    row[COL_SERVICE_CRITICALITY].value = c.service_criticality
                    row[COL_SERVICE_EVALUATION_STATUS].value = c.service_evaluation_status
                    row[COL_SERVICE_REASONING].value = c.service_reasoning
                    row[COL_SERVICE_MEASURES].value = c.service_measures

                    # CIA
                    row[COL_CONFIDENTIALITY].value = "x" if c.confidentiality else None
                    row[COL_INTEGRITY].value = "x" if c.integrity else None
                    row[COL_AVAILABILITY].value = "x" if c.availability else None
        
        # Delete all empty rows from bottom to top, else there will
        # be a problem when removing rows while iterating over them.
        for row in reversed(list(sheet.iter_rows(min_row=2))):
            if self.is_empty_row(row):
                sheet.delete_rows(row[0].row)

        self.doc_formulas.save(file_path)